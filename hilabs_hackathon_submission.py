import os
import sys
import email
import pandas as pd
from email import policy
import json
from pathlib import Path
from email.parser import BytesParser
from bs4 import BeautifulSoup
import openpyxl
from typing import List, Dict, Any
from transformers import AutoTokenizer, AutoModelForCausalLM, pipeline

# ------------ CONFIG ------------
MODEL_NAME = "microsoft/Phi-3-mini-4k-instruct"  # the tiny NuExtract model on HF
MAX_NEW_TOKENS = 512
# -------------------------------

# ---- Helper: load .eml and return text (plain + html cleaned)
def load_eml_text(path: Path) -> str:
    with open(path, "rb") as f:
        msg = BytesParser(policy=policy.default).parse(f)
    parts: List[str] = []

    def walk(m):
        if m.is_multipart():
            for sub in m.iter_parts():
                walk(sub)
        else:
            ctype = m.get_content_type()
            payload = m.get_content()
            if not payload:
                return
            if ctype == "text/plain":
                parts.append(str(payload))
            elif ctype == "text/html":
                soup = BeautifulSoup(str(payload), "lxml")
                parts.append(soup.get_text("\n", strip=True))

    walk(msg)
    text = "\n".join([p for p in parts if p and p.strip()])
    # tidy up whitespace
    text = "\n".join([ln.rstrip() for ln in text.splitlines()])
    return text.strip()

# ---- Load LLM pipeline (try GPU/auto, fallback to cpu)
def get_llm_pipe(model_name: str):
    try:
        tokenizer = AutoTokenizer.from_pretrained(model_name)
        model = AutoModelForCausalLM.from_pretrained(model_name, device_map="auto")
        gen = pipeline("text-generation", model=model, tokenizer=tokenizer, trust_remote_code=True)
        return gen
    except Exception:
        tokenizer = AutoTokenizer.from_pretrained(model_name)
        model = AutoModelForCausalLM.from_pretrained(model_name, device_map={"": "cpu"})
        gen = pipeline("text-generation", model=model, tokenizer=tokenizer, trust_remote_code=True)
        return gen

# ---- Build strict JSON prompt based on exact template fields ----
def make_prompt(email_text: str) -> str:
    # These are internal JSON keys the LLM must return — we'll map them to template headers later
    schema = {
        "transaction_type": "Add | Update | Term | Information not found",
        "transaction_attribute": "string or 'Information not found'",
        "effective_date": "MM/DD/YYYY or 'Information not found'",
        "term_date": "MM/DD/YYYY or 'Information not found'",
        "term_reason": "string or 'Information not found'",
        "provider_name": "string or 'Information not found' Only output the name not their designation",
        "provider_npi": "digits or 'Information not found'",
        "provider_specialty": "string or 'Information not found'",
        "state_license": "string or 'Information not found'",
        "organization_name": "string or 'Information not found'",
        "tin": "digits or 'Information not found'(It is the Tax Id No.)",
        "group_npi": "digits or 'Information not found' (It is NPI of Default Provider)",
        "complete_address": "string or 'Information not found'",
        "phone_number": "digits or 'Information not found'",
        "fax_number": "digits or 'Information not found'",
        "ppg_id": "string (single or comma-separated) or 'Information not found'",
        "line_of_business": "'Medicare' or 'Commercial' or 'Medical' or 'Information not found'  Only these strings should be the output"
    }

    schema_text = json.dumps(schema, indent=2)
    prompt = f"""
You are a STRUCTURED-EXTRACTION model. Extract values from the EMAIL below and RETURN STRICT JSON ONLY (no commentary).
If a value cannot be found, set it exactly to "Information not found".
Dates must be normalized to MM/DD/YYYY when possible.
Return a JSON object with the following keys and value formats (exact keys must be used):

{schema_text}

Email:
\"\"\"{email_text}\"\"\"

IMPORTANT: Return only valid JSON (a single JSON object) and nothing else.
"""
    return prompt

# ---- Call the LLM and parse JSON output ----
def extract_with_llm(pipe, email_text: str) -> Dict[str, Any]:
    prompt = make_prompt(email_text)
    out = pipe(prompt, max_new_tokens=MAX_NEW_TOKENS, do_sample=False, return_full_text=False)[0]["generated_text"]
    # Attempt to locate the JSON object inside output
    s = out.find("{")
    e = out.find("}")
    if s == -1 or e == -1 or e <= s:
        # model did not return JSON — return empty dict (we will fill "Information not found" later)
        return {}
    raw = out[s:e+1]
    try:
        data = json.loads(raw)
        return data if isinstance(data, dict) else {}
    except Exception:
        # best-effort: attempt small fix (replace single quotes)
        try:
            fixed = raw.replace("'", '"')
            return json.loads(fixed)
        except Exception:
            return {}

# ---- Mapping between template headers and JSON keys ----
TEMPLATE_TO_KEY = {
    "Transaction Type (Add/Update/Term)": "transaction_type",
    "Transaction Attribute": "transaction_attribute",
    "Effective Date": "effective_date",
    "Term Date": "term_date",
    "Term Reason": "term_reason",
    "Provider Name": "provider_name",
    "Provider NPI": "provider_npi",
    "Provider Specialty": "provider_specialty",
    "State License": "state_license",
    "Organization Name": "organization_name",
    "TIN": "tin",
    "Group NPI": "group_npi",
    "Complete Address": "complete_address",
    "Phone Number": "phone_number",
    "Fax Number": "fax_number",
    "PPG ID": "ppg_id",
    "Line Of Business (Medicare/Commercial/Medical)": "line_of_business",
}

# ---- Append row to template preserving header order
def append_row_to_template(template_path: Path, out_path: Path, row_values: List[Any]) -> None:
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    ws.delete_rows(0)
    ws.append(row_values)
    wb.save(out_path)

# ---- Main processing for one or many .eml files
def process_eml_files(eml_paths: List[Path], template_path: Path, out_path: Path, pipe):
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    # Validate headers against our mapping (if unknown header, we will write empty)
    keys_in_order = [TEMPLATE_TO_KEY.get(h, None) for h in headers]

    # Prepare output workbook by copying template once
    wb.save(out_path)  # initial copy 

    for eml_path in eml_paths:
        email_text = load_eml_text(eml_path)
        extracted = extract_with_llm(pipe, email_text)
        # Ensure all expected JSON keys exist; fill missing with "Information not found"
        final = {}
        for k in set(TEMPLATE_TO_KEY.values()):
            v = extracted.get(k) if isinstance(extracted, dict) else None
            if v is None or (isinstance(v, str) and v.strip() == ""):
                final[k] = "Information not found"
            else:
                # Convert lists to comma-separated strings
                if isinstance(v, list):
                    final[k] = ", ".join(str(x) for x in v)
                else:
                    final[k] = str(v).strip()

        # Build row in template order
        row = []
        for key in keys_in_order:
            if key is None:
                # unknown header present in template - write empty placeholder
                row.append("Information not found")
            else:
                row.append(final.get(key, "Information not found"))

        append_row_to_template(template_path, out_path, row)
        print(f"Processed: {eml_path}")

# ---- Main
# ---- Main
def main():
    import argparse

    parser = argparse.ArgumentParser(
        description="Parse .eml roster emails and export to standardized Excel format."
    )
    parser.add_argument(
        "eml_input",
        help="Path to a single .eml file OR a folder containing multiple .eml files."
    )
    parser.add_argument(
        "template",
        help="Path to the Excel template file (provided in hackathon dataset)."
    )
    parser.add_argument(
        "output",
        help="Path for the output Excel file."
    )
    parser.add_argument(
        "-v", "--verbose",
        action="store_true",
        help="Enable detailed logs for debugging."
    )
    parser.add_argument(
        "-b", "--batch",
        type=int,
        default=1,
        help="Number of emails to process in a batch (default=1)."
    )

    args = parser.parse_args()

    eml_input = Path(args.eml_input)
    template_path = Path(args.template)
    out_path = Path(args.output)

    # Collect .eml files
    if eml_input.is_dir():
        eml_paths = sorted(list(eml_input.glob("*.eml")))
    elif eml_input.is_file() and eml_input.suffix.lower() == ".eml":
        eml_paths = [eml_input]
    else:
        raise SystemExit("❌ Input must be an .eml file or a folder containing .eml files.")

    if not eml_paths:
        raise SystemExit("❌ No .eml files found in input.")

    if args.verbose:
        print(f"📂 Found {len(eml_paths)} .eml files to process.")

    # Load model/pipeline
    print(f"🔄 Loading LLM model: {MODEL_NAME} ...")
    pipe = get_llm_pipe(MODEL_NAME)
    print("✅ Model loaded. Beginning extraction...")

    # Process files in batches
    for i in range(0, len(eml_paths), args.batch):
        batch_files = eml_paths[i:i + args.batch]
        if args.verbose:
            print(f"⚡ Processing batch {i // args.batch + 1} with {len(batch_files)} file(s).")
        process_eml_files(batch_files, template_path, out_path, pipe)

    print(f"🎉 Done! Output saved at: {out_path.resolve()}")

if __name__ == "__main__":
    main()

